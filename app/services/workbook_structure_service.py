"""Workbook structure analysis (multi-sheet Excel) for robust Onyx workflows.

Goal: Avoid relying on a single 'first sheet' or hardcoded sheet names.
We scan all sheets, profile them, classify them, and recommend an analysis strategy.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.core.exceptions import InvalidInputError
from app.core.file_registry import file_registry
from app.utils.column_detector import detect_columns
from app.utils.dataframe_utils import normalize_column_name, safe_preview


@dataclass(frozen=True)
class SheetProfile:
    sheet_name: str
    row_count: int
    column_count: int
    non_empty_cell_ratio: float
    header_quality_score: float
    data_density_score: float
    text_column_count: int
    numeric_column_count: int
    date_column_count: int
    long_text_column_count: int
    likely_response_column_count: int
    likely_score_column_count: int
    duplicate_header_ratio: float
    empty_row_ratio: float
    formula_cell_ratio: float
    merged_cell_ratio: float
    unique_value_profile: dict[str, Any]
    sample_headers: list[str]
    sample_rows: list[dict]


class WorkbookStructureService:
    def analyze(self, file_id: str, task_hint: str = "") -> dict[str, Any]:
        meta = file_registry.get_meta(file_id)
        path = Path(meta.path)
        if not path.exists():
            raise InvalidInputError(f"Workbook bulunamadi: {meta.path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise InvalidInputError("analyze_workbook_structure sadece Excel workbook (.xlsx/.xlsm) icin kullanilir.")

        xls = pd.ExcelFile(path)
        sheet_names = list(xls.sheet_names)
        if not sheet_names:
            return {
                "file_id": file_id,
                "total_sheets": 0,
                "analyzable_sheets": [],
                "excluded_sheets": [],
                "recommended_strategy": "ask_user",
                "warnings": ["Workbook icinde sheet bulunamadi."],
            }

        # openpyxl diagnostics (merged/formula ratios)
        wb = load_workbook(path, read_only=True, data_only=False)

        profiles: list[dict[str, Any]] = []
        for sheet in sheet_names:
            prof = self._profile_sheet(path, wb, sheet)
            classification = self._classify_sheet(prof, task_hint=task_hint)
            profiles.append({**prof, **classification})

        analyzable = [p for p in profiles if p.get("is_analyzable")]
        excluded = [p for p in profiles if not p.get("is_analyzable")]

        recommended = self._recommend_strategy(analyzable)
        warnings: list[str] = []
        if not analyzable:
            warnings.append("Analiz edilebilir sheet bulunamadi; sheet yapisi belirsiz. inspect_workbook_sheets ile kontrol edip sheet_hint ile netlestirin.")
            recommended = "ask_user"
        else:
            # If top candidates are close, warn.
            ranked = sorted(analyzable, key=lambda p: float(p.get("confidence") or 0.0), reverse=True)
            if len(ranked) >= 2 and float(ranked[0]["confidence"]) - float(ranked[1]["confidence"]) < 0.08:
                warnings.append(
                    "Birden fazla olasi veri sheet'i var. Sistem en yuksek skorlu sheet(ler)i sececek; emin degilsen sheet adini belirt."
                )

        return {
            "file_id": file_id,
            "total_sheets": int(len(sheet_names)),
            "analyzable_sheets": [
                {
                    "sheet_name": p["sheet_name"],
                    "confidence": p["confidence"],
                    "reason": p["reason"],
                    "classification": p["classification"],
                    "profile": p["profile"],
                    "detected_columns": p["detected_columns"],
                }
                for p in sorted(analyzable, key=lambda x: float(x.get("confidence") or 0.0), reverse=True)
            ],
            "excluded_sheets": [
                {
                    "sheet_name": p["sheet_name"],
                    "confidence": p["confidence"],
                    "reason": p["reason"],
                    "classification": p["classification"],
                    "profile": p["profile"],
                    "detected_columns": p["detected_columns"],
                }
                for p in sorted(excluded, key=lambda x: float(x.get("confidence") or 0.0), reverse=True)
            ],
            "recommended_strategy": recommended,
            "warnings": warnings,
        }

    def _profile_sheet(self, path: Path, wb, sheet_name: str) -> dict[str, Any]:
        # Read a bounded preview to keep memory predictable.
        raw = pd.read_excel(path, sheet_name=sheet_name, nrows=400)
        # Keep a copy for empty-row ratio before dropping.
        raw_rows = int(len(raw))
        raw_cols = int(len(raw.columns))
        empty_row_ratio = float(raw.isna().all(axis=1).mean()) if raw_rows else 1.0

        df = raw.dropna(how="all").dropna(axis=1, how="all")
        row_count = int(len(df))
        column_count = int(len(df.columns))

        non_empty_cell_ratio = 0.0
        if row_count and column_count:
            non_empty_cell_ratio = float(df.notna().sum().sum() / (row_count * column_count))

        sample_headers = [str(c) for c in df.columns[:30]]
        sample_rows = safe_preview(df, 5) if row_count and column_count else []

        # Header quality: avoid Unnamed, empty headers, duplicates.
        header_norms = [normalize_column_name(c) for c in df.columns]
        non_empty_headers = [h for h in header_norms if h and not h.startswith("unnamed")]
        dup_ratio = 0.0
        if column_count:
            dup_ratio = float(pd.Series(header_norms).duplicated().mean())
        header_quality = 0.0
        if column_count:
            header_quality = float(len(non_empty_headers) / column_count) * (1.0 - min(1.0, dup_ratio))

        # Data density: similar to non_empty_cell_ratio but penalize sparse tables.
        data_density = non_empty_cell_ratio

        numeric_cols = list(df.select_dtypes(include="number").columns)
        date_cols = list(df.select_dtypes(include=["datetime", "datetimetz"]).columns)
        text_cols = [c for c in df.columns if c not in numeric_cols and c not in date_cols]

        long_text_cols = []
        for c in text_cols:
            s = df[c].dropna().astype(str)
            if s.empty:
                continue
            avg_len = float(s.head(200).map(len).mean())
            uniq_ratio = float(s.nunique(dropna=True) / max(len(s), 1))
            if avg_len >= 35 and uniq_ratio >= 0.25:
                long_text_cols.append(str(c))

        detected = detect_columns(df) if row_count and column_count else {
            "likely_score_columns": [],
            "group_columns": [],
            "comment_columns": [],
            "categorical_columns": [],
            "date_columns": [],
            "key_columns": [],
        }
        likely_score_count = int(len(detected.get("likely_score_columns") or []))
        likely_response_count = int(len(long_text_cols) + len(detected.get("comment_columns") or []))

        unique_profile: dict[str, Any] = {}
        for c in list(df.columns)[:8]:
            s = df[c]
            nn = int(s.notna().sum())
            uniq = int(s.nunique(dropna=True))
            unique_profile[str(c)] = {"non_null": nn, "unique": uniq, "unique_ratio": round(uniq / max(nn, 1), 4)}

        formula_ratio = 0.0
        merged_ratio = 0.0
        try:
            ws = wb[sheet_name]
            # merged cells
            merged_count = len(getattr(ws.merged_cells, "ranges", []) or [])
            # normalize by area of a small window to avoid weird max_row effects
            window_cells = max(1, min(int(ws.max_row), 200) * min(int(ws.max_column), 30))
            merged_ratio = float(min(1.0, merged_count / max(window_cells, 1)))

            # formula cells (sample window)
            formula_cells = 0
            scanned = 0
            for r in range(1, min(int(ws.max_row), 200) + 1):
                for c in range(1, min(int(ws.max_column), 30) + 1):
                    cell = ws.cell(row=r, column=c)
                    scanned += 1
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        formula_cells += 1
            formula_ratio = float(formula_cells / max(scanned, 1))
        except Exception:
            pass

        profile = SheetProfile(
            sheet_name=sheet_name,
            row_count=row_count,
            column_count=column_count,
            non_empty_cell_ratio=round(non_empty_cell_ratio, 4),
            header_quality_score=round(header_quality, 4),
            data_density_score=round(data_density, 4),
            text_column_count=int(len(text_cols)),
            numeric_column_count=int(len(numeric_cols)),
            date_column_count=int(len(date_cols)),
            long_text_column_count=int(len(long_text_cols)),
            likely_response_column_count=int(likely_response_count),
            likely_score_column_count=int(likely_score_count),
            duplicate_header_ratio=round(float(dup_ratio), 4),
            empty_row_ratio=round(float(empty_row_ratio), 4),
            formula_cell_ratio=round(float(formula_ratio), 4),
            merged_cell_ratio=round(float(merged_ratio), 4),
            unique_value_profile=unique_profile,
            sample_headers=sample_headers,
            sample_rows=sample_rows,
        )

        # Column buckets for consumers
        detected_columns = {
            "long_text_columns": long_text_cols,
            "numeric_score_columns": detected.get("likely_score_columns") or [],
            "category_columns": detected.get("categorical_columns") or [],
            "date_columns": detected.get("date_columns") or [],
        }

        return {
            "sheet_name": sheet_name,
            "profile": profile.__dict__,
            "detected_columns": detected_columns,
            "raw_preview_shape": {"rows": raw_rows, "cols": raw_cols},
        }

    def _classify_sheet(self, prof: dict[str, Any], task_hint: str = "") -> dict[str, Any]:
        p = prof["profile"]
        sheet_name = str(prof["sheet_name"])
        hint = normalize_column_name(task_hint)

        row_count = int(p["row_count"])
        col_count = int(p["column_count"])
        density = float(p["data_density_score"])
        header_q = float(p["header_quality_score"])
        formula_ratio = float(p["formula_cell_ratio"])

        # Data score heuristics
        score = 0.0
        reasons: list[str] = []

        if row_count >= 5:
            score += 0.18
            reasons.append("rows>=5")
        if row_count >= 20:
            score += 0.22
            reasons.append("rows>=20")
        if row_count >= 100:
            score += 0.12
            reasons.append("rows>=100")
        if col_count >= 2:
            score += 0.10
            reasons.append("cols>=2")
        if col_count >= 3:
            score += 0.06
            reasons.append("cols>=3")
        if 3 <= col_count <= 60:
            score += 0.12
            reasons.append("cols in range")
        if header_q >= 0.65:
            score += 0.18
            reasons.append("good headers")
        if density >= 0.25:
            score += 0.12
            reasons.append("dense enough")
        if int(p["likely_score_column_count"]) > 0 or int(p["long_text_column_count"]) > 0:
            score += 0.12
            reasons.append("analytic columns")

        # Penalties for summary-like sheets
        if row_count <= 2 or col_count <= 1:
            score -= 0.35
            reasons.append("too small")
        if formula_ratio >= 0.25 and row_count <= 60:
            score -= 0.15
            reasons.append("formula-heavy")
        if density < 0.08:
            score -= 0.15
            reasons.append("very sparse")

        # Task hint: open-ended survey/comment requests should favor long-text sheets.
        if any(tok in hint for tok in ("yorum", "comment", "feedback", "sikayet", "şikayet", "oner", "öneri", "open ended", "acik uclu", "açık uçlu")):
            if int(p["long_text_column_count"]) > 0 or int(p["likely_response_column_count"]) > 0:
                score += 0.12
                reasons.append("matches open-ended task")
            else:
                score -= 0.08
                reasons.append("no long-text for open-ended task")

        confidence = max(0.0, min(1.0, score))
        is_analyzable = confidence >= 0.55

        classification = "unknown"
        if is_analyzable:
            classification = "data_sheet"
        else:
            if row_count <= 6 and density <= 0.2:
                classification = "metadata_sheet"
            elif formula_ratio >= 0.25:
                classification = "summary_sheet"
            else:
                classification = "unknown"

        reason = f"{sheet_name}: " + ", ".join(reasons[:6]) if reasons else "heuristic"
        return {
            "classification": classification,
            "is_analyzable": bool(is_analyzable),
            "confidence": round(float(confidence), 4),
            "reason": reason,
        }

    @staticmethod
    def _recommend_strategy(analyzable: list[dict[str, Any]]) -> str:
        if not analyzable:
            return "ask_user"
        if len(analyzable) == 1:
            return "single_sheet"
        # Check schema similarity (header overlap) to decide merge vs separate.
        headers = []
        for p in analyzable:
            cols = set(normalize_column_name(c) for c in (p.get("profile", {}).get("sample_headers") or []))
            headers.append(cols)
        if not headers:
            return "ask_user"
        # average pairwise jaccard
        sims = []
        for i in range(len(headers)):
            for j in range(i + 1, len(headers)):
                a = headers[i]
                b = headers[j]
                if not a or not b:
                    continue
                inter = len(a & b)
                union = len(a | b)
                sims.append(inter / max(union, 1))
        avg = sum(sims) / len(sims) if sims else 0.0
        if avg >= 0.6:
            return "merge_similar_sheets"
        return "analyze_separately"


workbook_structure_service = WorkbookStructureService()
