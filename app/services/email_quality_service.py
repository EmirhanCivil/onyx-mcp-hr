"""Email quality audit — categorize invalid emails and export annotated xlsx."""

from __future__ import annotations

import re
from typing import Any

import pandas as pd

from app.core.exceptions import AppError
from app.core.file_registry import file_registry
from app.utils.file_utils import ensure_parent, make_job_id, safe_filename
from app.config import settings


# Validation rules (RFC 5322 simplified, practical bias)
_LOCAL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+$")
_DOMAIN_RE = re.compile(r"^[A-Za-z0-9.\-]+$")
_TLD_RE = re.compile(r"^[A-Za-z]{2,}$")


# Reason → (display label TR, severity)
_REASONS: dict[str, str] = {
    "valid": "Geçerli",
    "empty": "Boş",
    "no_at": "@ yok",
    "multiple_at": "Birden fazla @",
    "whitespace": "Boşluk içeriyor",
    "no_domain_dot": "Domain'de nokta yok",
    "short_tld": "TLD çok kısa (< 2)",
    "invalid_tld": "TLD geçersiz",
    "consecutive_dots": "Ardışık nokta (..)",
    "leading_or_trailing_dot": "Başta/sonda nokta",
    "invalid_local_chars": "Yerel kısımda geçersiz karakter",
    "invalid_domain_chars": "Domain kısmında geçersiz karakter",
    "empty_local": "Yerel kısım boş",
    "empty_domain": "Domain boş",
}


def _classify_email(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, float) and pd.isna(value):
        return "empty"
    s = str(value).strip()
    if not s:
        return "empty"
    if any(ch.isspace() for ch in s):
        return "whitespace"
    at_count = s.count("@")
    if at_count == 0:
        return "no_at"
    if at_count > 1:
        return "multiple_at"
    local, _, domain = s.partition("@")
    if not local:
        return "empty_local"
    if not domain:
        return "empty_domain"
    if local.startswith(".") or local.endswith(".") or domain.startswith(".") or domain.endswith("."):
        return "leading_or_trailing_dot"
    if ".." in local or ".." in domain:
        return "consecutive_dots"
    if not _LOCAL_RE.match(local):
        return "invalid_local_chars"
    if "." not in domain:
        return "no_domain_dot"
    if not _DOMAIN_RE.match(domain):
        return "invalid_domain_chars"
    tld = domain.rsplit(".", 1)[-1]
    if len(tld) < 2:
        return "short_tld"
    if not _TLD_RE.match(tld):
        return "invalid_tld"
    return "valid"


def _resolve_file(file_query: str, file_id: str) -> str:
    if file_id:
        return file_id
    if not file_query:
        return ""
    from app.services.excel_service import excel_service
    try:
        excel_service.scan_uploads()
    except Exception:
        pass
    selected = excel_service.select_file(file_query, "")
    if selected and "file_id" in selected:
        return selected["file_id"]
    return ""


def _detect_email_column(df: pd.DataFrame, hint: str) -> str:
    if hint and hint in df.columns:
        return hint
    candidates = [c for c in df.columns if isinstance(c, str)]
    # Prefer exact-ish
    priority = ["email", "e-mail", "e_mail", "mail", "eposta", "e-posta", "e_posta", "aday-email", "aday_email"]
    lowered = {c.strip().lower(): c for c in candidates}
    for p in priority:
        if p in lowered:
            return lowered[p]
    # Substring match
    for low, original in lowered.items():
        if "email" in low or "mail" in low or "eposta" in low or "e-posta" in low:
            return original
    return ""


class EmailQualityService:
    """Detect & categorize invalid emails; optionally export annotated xlsx."""

    def audit(
        self,
        file_id: str = "",
        file_query: str = "",
        email_column: str = "",
        export: bool = True,
        only_invalid_export: bool = False,
        sample_limit: int = 15,
    ) -> dict[str, Any]:
        resolved_id = _resolve_file(file_query, file_id)
        if not resolved_id:
            raise AppError(
                "FILE_NOT_FOUND",
                f"Dosya bulunamadı (file_id='{file_id}', file_query='{file_query}').",
            )

        df = file_registry.get_frame(resolved_id)
        if df is None or df.empty:
            raise AppError("EMPTY_FILE", "Dosya boş veya yüklenemedi.")

        meta = file_registry.get_meta(resolved_id)
        col = email_column or _detect_email_column(df, "")
        if not col:
            raise AppError(
                "EMAIL_COLUMN_MISSING",
                "Email kolonu otomatik tespit edilemedi. `email_column` parametresi ile elle belirt.",
            )

        # Classify each row
        reasons = df[col].apply(_classify_email)
        total = int(len(df))
        valid_mask = reasons == "valid"
        valid_count = int(valid_mask.sum())
        invalid_count = total - valid_count

        # Group counts
        by_reason: dict[str, int] = {}
        for r, c in reasons.value_counts().items():
            if r != "valid":
                by_reason[r] = int(c)

        # Sample invalid (per reason if possible)
        sample = []
        invalid_df = df.loc[~valid_mask, [col]].copy()
        invalid_df["_reason"] = reasons[~valid_mask].values
        for r in by_reason:
            picks = invalid_df.loc[invalid_df["_reason"] == r, col].head(3).tolist()
            for v in picks:
                sample.append({"value": "" if v is None else str(v),
                               "reason_code": r,
                               "reason": _REASONS.get(r, r)})
            if len(sample) >= sample_limit:
                break
        sample = sample[:sample_limit]

        # Build summary
        summary = {
            "total": total,
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "invalid_pct": round((invalid_count / total) * 100, 2) if total else 0.0,
            "email_column": col,
            "by_reason": [
                {"code": code, "label": _REASONS.get(code, code), "count": cnt}
                for code, cnt in sorted(by_reason.items(), key=lambda x: -x[1])
            ],
            "sample_invalid": sample,
        }

        result: dict[str, Any] = {
            "file_id": resolved_id,
            "file_name": meta.name if meta else "",
            "summary": summary,
        }

        files: list[str] = []
        outputs: list[dict[str, Any]] = []

        if export:
            export_path = self._export_annotated_xlsx(
                df,
                col,
                reasons,
                meta_name=(meta.name if meta else "email_audit"),
                only_invalid=only_invalid_export,
            )
            files.append(export_path)
            outputs.append({
                "type": "file",
                "title": export_path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1],
                "description": "Email kalite raporu (geçersiz satırlar kırmızı işaretli)",
                "format": "xlsx",
                "path": export_path,
            })

        result["files"] = files
        result["generated_outputs"] = outputs
        return result

    def _export_annotated_xlsx(
        self,
        df: pd.DataFrame,
        email_col: str,
        reasons: pd.Series,
        meta_name: str,
        only_invalid: bool,
    ) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        if only_invalid:
            mask = reasons != "valid"
            export_df = df.loc[mask].copy()
            export_reasons = reasons.loc[mask].copy()
        else:
            export_df = df.copy()
            export_reasons = reasons.copy()

        # Add status columns at the end
        export_df = export_df.reset_index(drop=True)
        export_reasons = export_reasons.reset_index(drop=True)
        export_df["Email Durum Kodu"] = export_reasons.values
        export_df["Email Durum"] = export_reasons.map(_REASONS).fillna("Bilinmeyen").values

        job = make_job_id("email_audit")
        base = safe_filename(meta_name.rsplit(".", 1)[0]) if "." in meta_name else safe_filename(meta_name)
        target = settings.EXPORT_DIR / job / f"{base}_email_audit.xlsx"
        ensure_parent(target)

        # Styles
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_fill = PatternFill("solid", fgColor="0F766E")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        text_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        invalid_fill = PatternFill("solid", fgColor="FEE2E2")  # rose-100
        invalid_email_fill = PatternFill("solid", fgColor="FCA5A5")  # rose-300 — emphasize email cell
        valid_email_fill = PatternFill("solid", fgColor="DCFCE7")  # green-100
        thin = Side(style="thin", color="94A3B8")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        invalid_font = Font(color="991B1B", name="Calibri", size=10.5, bold=False)
        valid_font = Font(color="166534", name="Calibri", size=10.5)

        wb = Workbook()
        ws = wb.active
        ws.title = "Email Audit"

        # Header row
        cols = list(export_df.columns)
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border
        ws.row_dimensions[1].height = 26

        email_col_idx = cols.index(email_col) + 1 if email_col in cols else None
        status_col_idx = cols.index("Email Durum") + 1
        status_code_col_idx = cols.index("Email Durum Kodu") + 1

        max_len = {c: len(str(c)) for c in cols}

        for r_offset, (_, row) in enumerate(export_df.iterrows(), start=2):
            reason_code = row["Email Durum Kodu"]
            is_invalid = reason_code != "valid"
            for c_idx, col in enumerate(cols, start=1):
                value = row[col]
                if isinstance(value, float) and pd.isna(value):
                    value = ""
                cell = ws.cell(row=r_offset, column=c_idx, value=value)
                cell.alignment = text_align
                cell.border = cell_border
                if is_invalid:
                    cell.fill = invalid_fill
                    if c_idx == email_col_idx:
                        cell.fill = invalid_email_fill
                        cell.font = invalid_font
                    elif c_idx == status_col_idx or c_idx == status_code_col_idx:
                        cell.font = invalid_font
                else:
                    if c_idx == email_col_idx:
                        cell.fill = valid_email_fill
                    if c_idx == status_col_idx or c_idx == status_code_col_idx:
                        cell.fill = valid_email_fill
                        cell.font = valid_font
                if value not in (None, ""):
                    max_len[col] = max(max_len[col], min(60, len(str(value))))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

        for c_idx, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(48, max_len[col] + 4))

        wb.save(target)
        return str(target.resolve())


email_quality_service = EmailQualityService()
